"""Field collection Excel: export a fillable sheet and import it back (round-trip)."""
from __future__ import annotations

import io
from typing import Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .state import new_id

TEAL = "FF007E7A"
LIGHT = "FFE6F2F1"

# Fixed layout so import can locate values deterministically.
BASIC_ROWS = {
    3: "atividade",
    4: "aplicadores",
    5: "data_analise",
    6: "area",
    7: "gerencia",
    8: "supervisao",
    9: "revisao",
    10: "data_revisao",
    11: "section_label",
}
BASIC_LABELS = {
    "atividade": "Atividade em análise",
    "aplicadores": "Aplicadores",
    "data_analise": "Data da análise",
    "area": "Área",
    "gerencia": "Gerência",
    "supervisao": "Supervisão",
    "revisao": "Revisão",
    "data_revisao": "Data da revisão",
    "section_label": "Rótulo do grupo",
}
TASK_HEADER_ROW = 13
TASK_FIRST_ROW = 14
TASK_COLS = ["Tarefa", "Task", "Descrição", "Início", "Fim", "I x E (interna/externa)"]


def _hfont(white=True, bold=True):
    return Font(name="HELVETICA", size=10, bold=bold, color="FFFFFFFF" if white else "FF0F172A")


def build_bytes(project: dict) -> bytes:
    basic = project.get("basic", {})
    wb = Workbook()
    ws = wb.active
    ws.title = "Coleta"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 34
    for col in "CDEF":
        ws.column_dimensions[col].width = 16

    ws.merge_cells("A1:F1")
    ws["A1"] = "SMED - Coleta de campo"
    ws["A1"].font = _hfont()
    ws["A1"].fill = PatternFill("solid", fgColor=TEAL)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    for row, key in BASIC_ROWS.items():
        ws[f"A{row}"] = BASIC_LABELS[key]
        ws[f"A{row}"].font = Font(name="HELVETICA", size=10, bold=True)
        if key == "section_label":
            ws[f"B{row}"] = project.get("section_label", "")
        else:
            ws[f"B{row}"] = basic.get(key, "")

    for idx, label in enumerate(TASK_COLS):
        cell = ws.cell(row=TASK_HEADER_ROW, column=idx + 1, value=label)
        cell.font = _hfont()
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    analysis = project.get("analysis", {})
    r = TASK_FIRST_ROW
    for task in project.get("tasks", []):
        ws.cell(row=r, column=1, value=task.get("tarefa", ""))
        ws.cell(row=r, column=2, value=task.get("task", ""))
        ws.cell(row=r, column=3, value=task.get("descricao", ""))
        ws.cell(row=r, column=4, value=task.get("inicio", ""))
        ws.cell(row=r, column=5, value=task.get("fim", ""))
        ws.cell(row=r, column=6, value=(task.get("ie_inicial") or ""))
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_bytes(data: bytes) -> Tuple[dict, str]:
    """Parse a field Excel back into partial project data. Returns (partial, error)."""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:  # noqa: BLE001 - surface any openpyxl error to the user
        return {}, f"Não foi possível ler o arquivo: {exc}"
    ws = wb.active

    basic = {}
    section_label = ""
    for row, key in BASIC_ROWS.items():
        value = ws[f"B{row}"].value
        value = "" if value is None else str(value).strip()
        if key == "section_label":
            section_label = value
        else:
            basic[key] = value

    tasks = []
    analysis = {}
    r = TASK_FIRST_ROW
    while True:
        tarefa = ws.cell(row=r, column=1).value
        descricao = ws.cell(row=r, column=3).value
        inicio = ws.cell(row=r, column=4).value
        fim = ws.cell(row=r, column=5).value
        if all(v in (None, "") for v in (tarefa, descricao, inicio, fim)):
            break
        tid = new_id()
        ie = ws.cell(row=r, column=6).value
        ie = "" if ie is None else str(ie).strip().lower()
        ie = ie if ie in ("interna", "externa") else ""
        tasks.append({
            "id": tid,
            "tarefa": "" if tarefa is None else str(tarefa),
            "task": "" if ws.cell(row=r, column=2).value is None else str(ws.cell(row=r, column=2).value),
            "descricao": "" if descricao is None else str(descricao),
            "inicio": "" if inicio is None else str(inicio),
            "fim": "" if fim is None else str(fim),
            "ie_inicial": ie,
        })
        if ie:
            analysis[tid] = {"ie": ie, "e": False, "c": False, "r": False,
                             "s": False, "ganho": 0, "kaizen": "", "o_que_e": ""}
        r += 1
        if r > 5000:  # safety bound
            break

    return {
        "basic": basic,
        "section_label": section_label,
        "tasks": tasks,
        "analysis": analysis,
    }, ""
