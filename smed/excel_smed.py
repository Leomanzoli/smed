"""Generate the SMED form Excel, matching assets/Formulario de Aplicacao do SMED.xlsx."""
from __future__ import annotations

import io
import re
from datetime import date, datetime, time, timedelta
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from .compute import compute_row
from .i18n import get_lang

TEAL = "FF007E7A"
LIGHT = "FFE6F2F1"

# Localized labels for the SMED form (falls back to PT).
_L = {
    "pt": {
        "title": "SMED - Single Minute Exchange of Die",
        "atividade": "Atividade: {}", "data_analise": "Data análise:",
        "elaborador": "Elaborador:", "revisao": "Revisão:", "data_revisao": "Data da revisão:",
        "tarefa": "Tarefa", "task": "Task", "descricao": "Descrição",
        "inicio": "Início", "fim": "Fim", "tempo": "Tempo",
        "ie": "Análise I x E", "interna": "Interna", "externa": "Externa",
        "interno": "Interno", "externo": "Externo", "ecrs": "Análise ECRS",
        "ganho": "Ganho estimado", "tempo_final": "Tempo Final",
        "melhoria": "Melhoria - Kaizens necessários", "qual": "Qual?", "oque": "O que é?",
        "total": "Total", "reducao": "Redução total:",
    },
    "en": {
        "title": "SMED - Single Minute Exchange of Die",
        "atividade": "Activity: {}", "data_analise": "Analysis date:",
        "elaborador": "Author:", "revisao": "Revision:", "data_revisao": "Revision date:",
        "tarefa": "Task", "task": "No.", "descricao": "Description",
        "inicio": "Start", "fim": "End", "tempo": "Time",
        "ie": "I x E analysis", "interna": "Internal", "externa": "External",
        "interno": "Internal", "externo": "External", "ecrs": "ECRS analysis",
        "ganho": "Estimated gain", "tempo_final": "Final time",
        "melhoria": "Improvement - Kaizens needed", "qual": "Which?", "oque": "What is it?",
        "total": "Total", "reducao": "Total reduction:",
    },
}

FMT_CLOCK = "hh:mm"
FMT_DUR = "[h]:mm:ss"
FMT_TOTAL = "[h]:mm:ss;@"
FMT_PCT = "0%"
FMT_DATE = "dd/mm/yyyy"

FONT_NAME = "HELVETICA"
_thin = Side(style="thin", color="FF9CA3AF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

MIN_ROWS = 12


def _font(bold=False, white=False, size=10):
    color = "FFFFFFFF" if white else "FF0F172A"
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)


def _fill(color):
    return PatternFill("solid", fgColor=color)


def _align(h="center", wrap=True):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def _set(ws, coord, value=None, font=None, fill=None, align=None, fmt=None, border=True):
    cell = ws[coord]
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = BORDER
    return cell


def _parse_date(value) -> Optional[date]:
    if not value:
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_time(value) -> Optional[time]:
    if not value:
        return None
    if isinstance(value, time):
        return value
    s = str(value).strip().replace("h", ":")
    parts = s.split(":")
    try:
        h = int(parts[0])
        m = int(parts[1]) if len(parts) > 1 and parts[1] != "" else 0
    except (ValueError, IndexError):
        return None
    if 0 <= h <= 23 and 0 <= m <= 59:
        return time(h, m)
    return None


def _safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\[\]:*?/\\]", " ", (name or "SMED")).strip() or "SMED"
    return name[:31]


def build_workbook(project: dict) -> Workbook:
    basic = project.get("basic", {})
    tasks = project.get("tasks", [])
    analysis = project.get("analysis", {})
    tr = _L.get(get_lang(), _L["pt"])

    section = project.get("section_label") or basic.get("atividade") or "SMED"
    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(project.get("name") or section or "SMED")
    ws.sheet_view.showGridLines = False

    # Column widths (A spacer .. T spacer)
    widths = {
        "A": 2, "B": 16, "C": 6, "D": 34, "E": 8, "F": 8, "G": 11,
        "H": 8, "I": 8, "J": 9, "K": 9, "L": 4, "M": 4, "N": 4, "O": 4,
        "P": 12, "Q": 11, "R": 16, "S": 26, "T": 2,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # --- Title B2:S2 ---
    ws.merge_cells("B2:S2")
    _set(ws, "B2", tr["title"],
         font=_font(bold=True, white=True, size=13), fill=_fill(TEAL), align=_align())
    for col in "CDEFGHIJKLMNOPQRS":
        _set(ws, f"{col}2", fill=_fill(TEAL), align=_align())
    ws.row_dimensions[2].height = 26

    # --- Basic info rows 4-5 ---
    ws.merge_cells("B4:G5")
    _set(ws, "B4", tr["atividade"].format(basic.get('atividade', '')),
         font=_font(bold=True), fill=_fill(LIGHT), align=_align("left"))
    for coord in ("C4", "D4", "E4", "F4", "G4", "B5", "C5", "D5", "E5", "F5", "G5"):
        _set(ws, coord, fill=_fill(LIGHT))

    ws.merge_cells("H4:I4")
    _set(ws, "H4", tr["data_analise"], font=_font(bold=True), align=_align("left"))
    ws.merge_cells("H5:I5")
    d = _parse_date(basic.get("data_analise"))
    _set(ws, "H5", d or basic.get("data_analise", ""), font=_font(), align=_align("left"),
         fmt=FMT_DATE if d else None)
    _set(ws, "I4"); _set(ws, "I5")

    ws.merge_cells("J4:P4")
    _set(ws, "J4", tr["elaborador"], font=_font(bold=True), align=_align("left"))
    ws.merge_cells("J5:P5")
    _set(ws, "J5", basic.get("aplicadores", ""), font=_font(), align=_align("left"))
    for col in "KLMNOP":
        _set(ws, f"{col}4"); _set(ws, f"{col}5")

    ws.merge_cells("Q4:R4")
    _set(ws, "Q4", tr["revisao"], font=_font(bold=True), align=_align("left"))
    ws.merge_cells("Q5:R5")
    _set(ws, "Q5", basic.get("revisao", ""), font=_font(), align=_align("left"))
    _set(ws, "R4"); _set(ws, "R5")
    _set(ws, "S4", tr["data_revisao"], font=_font(bold=True), align=_align("left"))
    dr = _parse_date(basic.get("data_revisao"))
    _set(ws, "S5", dr or basic.get("data_revisao", ""), font=_font(), align=_align("left"),
         fmt=FMT_DATE if dr else None)

    # --- Header rows 7-8 ---
    hf = _font(bold=True, white=True)
    def hcell(coord, text=None):
        _set(ws, coord, text, font=hf, fill=_fill(TEAL), align=_align())

    ws.merge_cells("B7:B8"); hcell("B7", tr["tarefa"])
    ws.merge_cells("C7:C8"); hcell("C7", tr["task"])
    ws.merge_cells("D7:G7"); hcell("D7", str(section).upper())
    for col in "EFG":
        hcell(f"{col}7")
    hcell("D8", tr["descricao"]); hcell("E8", tr["inicio"]); hcell("F8", tr["fim"]); hcell("G8", tr["tempo"])
    ws.merge_cells("H7:I7"); hcell("H7", tr["ie"]); hcell("I7")
    hcell("H8", tr["interna"]); hcell("I8", tr["externa"])
    ws.merge_cells("J7:K7"); hcell("J7", tr["tempo"]); hcell("K7")
    hcell("J8", tr["interno"]); hcell("K8", tr["externo"])
    ws.merge_cells("L7:O7"); hcell("L7", tr["ecrs"])
    for col in "MNO":
        hcell(f"{col}7")
    hcell("L8", "E"); hcell("M8", "C"); hcell("N8", "R"); hcell("O8", "S")
    ws.merge_cells("P7:P8"); hcell("P7", tr["ganho"])
    ws.merge_cells("Q7:Q8"); hcell("Q7", tr["tempo_final"])
    ws.merge_cells("R7:S7"); hcell("R7", tr["melhoria"]); hcell("S7")
    hcell("R8", tr["qual"]); hcell("S8", tr["oque"])
    ws.row_dimensions[7].height = 20
    ws.row_dimensions[8].height = 18

    # --- Data rows ---
    n = max(len(tasks), MIN_ROWS)
    first = 9
    last = first + n - 1
    for i in range(n):
        r = first + i
        task = tasks[i] if i < len(tasks) else {}
        a = analysis.get(task.get("id"), {}) if task else {}
        row = compute_row(task, a) if task else None

        _set(ws, f"B{r}", task.get("tarefa", "") if task else None, font=_font(), align=_align("left"))
        _set(ws, f"C{r}", task.get("task", "") if task else None, font=_font(), align=_align())
        _set(ws, f"D{r}", task.get("descricao", "") if task else None, font=_font(), align=_align("left"))

        ti = _parse_time(task.get("inicio")) if task else None
        tf = _parse_time(task.get("fim")) if task else None
        _set(ws, f"E{r}", ti, font=_font(), align=_align(), fmt=FMT_CLOCK if ti else None)
        _set(ws, f"F{r}", tf, font=_font(), align=_align(), fmt=FMT_CLOCK if tf else None)
        _set(ws, f"G{r}",
             f'=IF(OR(E{r}="",F{r}=""),"",F{r}-E{r}+IF(F{r}<E{r},1,0))',
             font=_font(), align=_align(), fmt=FMT_DUR)

        ie = (a.get("ie") or "").lower()
        _set(ws, f"H{r}", "X" if ie == "interna" else None, font=_font(), align=_align())
        _set(ws, f"I{r}", "X" if ie == "externa" else None, font=_font(), align=_align())
        _set(ws, f"J{r}", f'=IF(AND($H{r}="X",$G{r}<>""),$G{r},0)', font=_font(), align=_align(), fmt=FMT_DUR)
        _set(ws, f"K{r}", f'=IF(AND($I{r}="X",$G{r}<>""),$G{r},0)', font=_font(), align=_align(), fmt=FMT_DUR)

        _set(ws, f"L{r}", "X" if a.get("e") else None, font=_font(), align=_align())
        _set(ws, f"M{r}", "X" if a.get("c") else None, font=_font(), align=_align())
        _set(ws, f"N{r}", "X" if a.get("r") else None, font=_font(), align=_align())
        _set(ws, f"O{r}", "X" if a.get("s") else None, font=_font(), align=_align())

        ganho = row["ganho"] if row else 0
        _set(ws, f"P{r}", timedelta(minutes=ganho) if ganho else None,
             font=_font(), align=_align(), fmt=FMT_DUR)
        _set(ws, f"Q{r}", f'=IF($G{r}="","",$G{r}-IF($P{r}="",0,$P{r}))',
             font=_font(), align=_align(), fmt=FMT_DUR)
        _set(ws, f"R{r}", a.get("kaizen", "") if task else None, font=_font(), align=_align("left"))
        _set(ws, f"S{r}", a.get("o_que_e", "") if task else None, font=_font(), align=_align("left"))

    # Data validations (X list) for I/E and ECRS
    dv = DataValidation(type="list", formula1='"X,x"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"H{first}:I{last}")
    dv.add(f"L{first}:O{last}")

    # --- Totals row ---
    tot = last + 1
    _set(ws, f"B{tot}", tr["total"], font=_font(bold=True), fill=_fill(LIGHT), align=_align())
    for col in "CDEF":
        _set(ws, f"{col}{tot}", fill=_fill(LIGHT))
    _set(ws, f"G{tot}", f"=SUBTOTAL(9,G{first}:G{last})", font=_font(bold=True),
         fill=_fill(LIGHT), align=_align(), fmt=FMT_TOTAL)
    _set(ws, f"H{tot}", fill=_fill(LIGHT)); _set(ws, f"I{tot}", fill=_fill(LIGHT))
    _set(ws, f"J{tot}", f"=SUBTOTAL(9,J{first}:J{last})", font=_font(bold=True),
         fill=_fill(LIGHT), align=_align(), fmt=FMT_TOTAL)
    _set(ws, f"K{tot}", f"=SUBTOTAL(9,K{first}:K{last})", font=_font(bold=True),
         fill=_fill(LIGHT), align=_align(), fmt=FMT_TOTAL)
    for col in "LMNO":
        _set(ws, f"{col}{tot}", fill=_fill(LIGHT))
    _set(ws, f"P{tot}", f"=SUM(P{first}:P{last})", font=_font(bold=True),
         fill=_fill(LIGHT), align=_align(), fmt=FMT_TOTAL)
    _set(ws, f"Q{tot}", f"=SUM(Q{first}:Q{last})", font=_font(bold=True),
         fill=_fill(LIGHT), align=_align(), fmt=FMT_TOTAL)
    _set(ws, f"R{tot}", fill=_fill(LIGHT)); _set(ws, f"S{tot}", fill=_fill(LIGHT))

    # --- Reduction % row ---
    pct = tot + 1
    ws.merge_cells(f"N{pct}:P{pct}")
    _set(ws, f"N{pct}", tr["reducao"], font=_font(bold=True), fill=_fill(LIGHT), align=_align("right"))
    for col in "OP":
        _set(ws, f"{col}{pct}", fill=_fill(LIGHT))
    _set(ws, f"Q{pct}", f'=IF(G{tot}=0,0,1-(Q{tot}/G{tot}))', font=_font(bold=True),
         fill=_fill(LIGHT), align=_align(), fmt=FMT_PCT)

    ws.freeze_panes = "A9"
    ws.sheet_view.zoomScale = 100
    return wb


def build_bytes(project: dict) -> bytes:
    buf = io.BytesIO()
    build_workbook(project).save(buf)
    return buf.getvalue()
