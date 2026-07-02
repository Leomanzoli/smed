"""5W2H action plan Excel export."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .i18n import get_lang

TEAL = "FF007E7A"

# Column keys (order preserved) and localized headers.
COLUMN_KEYS = [
    "o_que", "por_que", "onde", "quando", "quem", "como", "quanto", "matricula", "email"]
_L = {
    "pt": {
        "title": "Plano de Ação 5W2H",
        "headers": ["O quê?", "Por quê?", "Onde?", "Quando?", "Quem?", "Como?",
                    "Quanto?", "Matrícula", "E-mail"],
    },
    "en": {
        "title": "5W2H Action Plan",
        "headers": ["What?", "Why?", "Where?", "When?", "Who?", "How?",
                    "How much?", "ID", "E-mail"],
    },
}


def build_bytes(project: dict) -> bytes:
    tr = _L.get(get_lang(), _L["pt"])
    columns = list(zip(COLUMN_KEYS, tr["headers"]))
    wb = Workbook()
    ws = wb.active
    ws.title = "5W2H"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title = ws.cell(row=1, column=1, value=tr["title"])
    title.font = Font(name="HELVETICA", size=13, bold=True, color="FFFFFFFF")
    title.fill = PatternFill("solid", fgColor=TEAL)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    header_row = 3
    for idx, (_key, label) in enumerate(columns):
        cell = ws.cell(row=header_row, column=idx + 1, value=label)
        cell.font = Font(name="HELVETICA", size=10, bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        letter = cell.column_letter
        ws.column_dimensions[letter].width = 26 if idx < 7 else 16

    r = header_row + 1
    for item in project.get("action_plan", []):
        for idx, (key, _label) in enumerate(columns):
            c = ws.cell(row=r, column=idx + 1, value=item.get(key, ""))
            c.font = Font(name="HELVETICA", size=10)
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        r += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
